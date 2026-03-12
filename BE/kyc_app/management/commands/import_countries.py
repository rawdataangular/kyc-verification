import openpyxl
from django.core.management.base import BaseCommand
from kyc_app.models import CountryMaster

class Command(BaseCommand):
    help = 'Import countries from Excel file'

    def handle(self, *args, **options):
        file_path = 'county list.xlsx'
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            # Mapping based on headers: ['COUNTRY', 'COUNTRY CODE', 'CODES 3Letter', 'CODES 2Letter']
            
            count = 0
            for row in ws.iter_rows(min_row=2):
                name = row[0].value
                dial_code = str(row[1].value).strip() if row[1].value else ""
                code_3 = str(row[2].value).strip() if row[2].value else ""
                code_2 = str(row[3].value).strip() if row[3].value else ""

                if not name:
                    continue
                
                # Update or create the country record
                obj, created = CountryMaster.objects.update_or_create(
                    name=name.strip(),
                    defaults={
                        'dial_code': dial_code,
                        'code_3': code_3,
                        'code_2': code_2,
                    }
                )
                if created:
                    count += 1
            
            self.stdout.write(self.style.SUCCESS(f'Successfully imported {count} new countries.'))
            self.stdout.write(self.style.SUCCESS('All country data updated.'))
            
        except Exception as e:
            self.stderr.write(self.style.ERROR(f'Error importing excel: {str(e)}'))
